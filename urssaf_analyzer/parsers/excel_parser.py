"""Parseur pour les fichiers Excel (bulletins de paie, livres de paie, exports comptables).

Detection automatique de la ligne d'en-tete, mapping des colonnes par mots-cles,
extraction multi-salaries avec deduplication par NIR ou nom+prenom.
"""

import re as _re
from decimal import Decimal
from pathlib import Path
from typing import Any

from urssaf_analyzer.core.exceptions import ParseError
from urssaf_analyzer.models.documents import (
    Document, Declaration, Cotisation, Employe, DateRange,
)
from urssaf_analyzer.config.constants import ContributionType
from urssaf_analyzer.parsers.base_parser import BaseParser
from urssaf_analyzer.utils.number_utils import parser_montant

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Mots indiquant une ligne de total/sous-total a ignorer
_TOTAL_KEYWORDS = {
    "total", "sous-total", "sous_total", "sous total",
    "totaux", "cumul", "cumuls", "recap", "recapitulatif",
    "general", "g\u00e9n\u00e9ral", "s/total", "net a payer global",
}

# ============================================================
# COLUMN MAPPING CONFIGURATION
# Each logical field -> list of possible normalized header keywords.
# Matching: exact match first, then "keyword IN header" (never reverse).
# ============================================================

_COLUMN_KEYWORDS = {
    "nom": [
        "nom", "nom_salarie", "nom_du_salarie", "nom_employe",
        "nom_de_l_employe", "nom_agent", "nom_famille", "patronyme",
    ],
    "prenom": [
        "prenom", "prenom_salarie", "prenom_du_salarie", "prenom_employe",
    ],
    "nom_prenom": [
        "nom_prenom", "nom_et_prenom", "prenom_nom", "identite",
        "salarie_nom_prenom", "nom_complet", "salarie", "employe",
        "intitule_salarie", "agent", "designation_salarie",
    ],
    "nir": [
        "nir", "numero_ss", "securite_sociale", "n_ss", "nss",
        "numero_securite_sociale", "n_securite_sociale",
    ],
    "matricule": [
        "matricule", "num_matricule", "numero_matricule", "mat",
        "n_mat", "id_salarie", "id_employe", "n_salarie",
    ],
    "base_brute": [
        "base_brute", "salaire_brut", "brut", "base",
        "remuneration_brute", "remuneration", "montant_brut",
        "sal_brut", "total_brut", "brut_mensuel",
        "brut_total", "salaire", "remuneration_totale",
    ],
    "net": [
        "net", "net_a_payer", "salaire_net", "montant_net",
        "net_paye", "net_verse", "net_mensuel", "a_payer",
    ],
    "taux_patronal": [
        "taux_patronal", "taux_employeur", "tx_patronal",
        "taux_part_employeur",
    ],
    "taux_salarial": [
        "taux_salarial", "taux_salarie", "tx_salarial",
        "taux_part_salarie",
    ],
    "montant_patronal": [
        "montant_patronal", "cotisation_employeur",
        "part_employeur", "charges_patronales",
        "cotisations_patronales", "patronal",
        "total_patronal",
    ],
    "montant_salarial": [
        "montant_salarial", "cotisation_salarie",
        "part_salarie", "charges_salariales",
        "cotisations_salariales", "retenues", "salarial",
        "total_salarial", "part_salariale",
    ],
    "type_cotisation": [
        "type_cotisation", "code_cotisation", "libelle_cotisation",
        "nature_cotisation", "rubrique",
    ],
    "heures": [
        "heures", "heures_travaillees", "nb_heures", "h_travaillees",
        "horaire", "heures_mensuelles",
    ],
    "statut": [
        "statut", "categorie", "classification", "coefficient",
    ],
    "poste": [
        "poste", "emploi", "fonction", "qualification", "metier",
    ],
    "service": [
        "service", "departement", "direction", "unite",
    ],
    "date_entree": [
        "date_entree", "date_embauche", "date_debut", "anciennete",
    ],
}


class ExcelParser(BaseParser):
    """Parse les fichiers Excel (.xlsx)."""

    def peut_traiter(self, chemin: Path) -> bool:
        return chemin.suffix.lower() in (".xlsx", ".xls")

    def extraire_metadata(self, chemin: Path) -> dict[str, Any]:
        if not HAS_OPENPYXL:
            return {"erreur": "openpyxl non installe"}
        try:
            wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
            metadata = {
                "format": "excel",
                "feuilles": wb.sheetnames,
                "nb_feuilles": len(wb.sheetnames),
            }
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                metadata[f"feuille_{sheet_name}_dims"] = ws.dimensions
            wb.close()
            return metadata
        except Exception as e:
            return {"format": "excel", "erreur": str(e)}

    def parser(self, chemin: Path, document: Document) -> list[Declaration]:
        if not HAS_OPENPYXL:
            raise ParseError("openpyxl n'est pas installe. Installer avec: pip install openpyxl")

        try:
            wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
        except Exception as e:
            raise ParseError(f"Impossible de lire le fichier Excel {chemin}: {e}") from e

        declarations = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            decl = self._parser_feuille(ws, sheet_name, document)
            if decl and (decl.cotisations or decl.employes):
                declarations.append(decl)

        wb.close()
        return declarations

    def _parser_feuille(
        self, ws: Any, nom_feuille: str, document: Document
    ) -> Declaration | None:
        """Parse une feuille Excel."""
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return None

        header_idx, header = self._trouver_entete(rows)
        if header_idx is None:
            return None

        col_map = self._mapper_colonnes(header)
        if not col_map:
            return None

        parse_log = []
        parse_log.append(f"Feuille: {nom_feuille}, en-tetes ligne {header_idx + 1}: {header}")
        parse_log.append(f"Colonnes mappees: {col_map}")

        cotisations = []
        employes_vus = {}   # cle de dedup -> Employe
        employe_ids = {}    # cle de dedup -> employe.id
        employes_meta = {}  # employe.id -> {heures, poste, service, ...}
        net_par_emp = {}    # employe.id -> net value

        for row_idx, row_data in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            if not row_data or all(c is None for c in row_data):
                continue
            if self._est_ligne_total(row_data):
                parse_log.append(f"  Ligne {row_idx}: ignoree (total/sous-total)")
                continue

            # Build mapped_row using col_map (logical field -> value)
            mapped_row = {}
            for field, col_idx in col_map.items():
                if col_idx < len(row_data):
                    val = row_data[col_idx]
                    if val is not None and str(val).strip():
                        mapped_row[field] = val

            if not mapped_row:
                continue

            # Extract employee
            employe = self._extraire_employe_mapped(mapped_row, document.id)
            emp_key = None
            if employe:
                if employe.nir:
                    emp_key = employe.nir
                elif employe.nom:
                    emp_key = f"{(employe.nom or '').strip().lower()}|{(employe.prenom or '').strip().lower()}"
                if emp_key and emp_key not in employes_vus:
                    employes_vus[emp_key] = employe
                    employe_ids[emp_key] = employe.id
                    parse_log.append(f"  Employe ligne {row_idx}: {employe.nom} {employe.prenom} NIR={employe.nir} key={emp_key}")
                    # Store extra metadata
                    meta = {}
                    for mf in ("heures", "poste", "service", "statut", "date_entree"):
                        if mapped_row.get(mf):
                            meta[mf] = str(mapped_row[mf]).strip()
                    if meta:
                        employes_meta[employe.id] = meta

            # Extract cotisation
            cotisation = self._extraire_cotisation_mapped(mapped_row, document.id)
            if cotisation:
                if emp_key and emp_key in employe_ids:
                    cotisation.employe_id = employe_ids[emp_key]
                cotisations.append(cotisation)

            # Store net per employee
            net_raw = mapped_row.get("net")
            if net_raw and emp_key and emp_key in employe_ids:
                net_par_emp[employe_ids[emp_key]] = self._to_decimal(net_raw)

        # Detect type: livre de paie or bulletin
        is_livre_paie = len(employes_vus) > 1
        type_decl = "livre_de_paie" if is_livre_paie else "bulletin"

        parse_log.append(f"Resultat: {len(employes_vus)} employes, {len(cotisations)} cotisations, type={type_decl}")

        # Calculate masse salariale per employee (max brut per employee, not sum)
        masse = Decimal("0")
        if employes_vus and cotisations:
            bruts_par_emp = {}
            for c in cotisations:
                eid = c.employe_id or "_global"
                if eid not in bruts_par_emp or c.base_brute > bruts_par_emp[eid]:
                    bruts_par_emp[eid] = c.base_brute
            masse = sum(bruts_par_emp.values())
        elif cotisations:
            bruts_uniques = set()
            for c in cotisations:
                if c.base_brute > 0:
                    bruts_uniques.add(c.base_brute)
            masse = sum(bruts_uniques) if bruts_uniques else sum(c.base_brute for c in cotisations)

        declaration = Declaration(
            type_declaration=type_decl,
            reference=document.nom_fichier or nom_feuille,
            cotisations=cotisations,
            employes=list(employes_vus.values()),
            effectif_declare=len(employes_vus),
            source_document_id=document.id,
            metadata={
                "type_document": "livre_de_paie" if is_livre_paie else "bulletin_de_paie",
                "parse_log": parse_log,
                "employes_meta": employes_meta,
                "net_par_employe": {k: float(v) for k, v in net_par_emp.items()},
            },
        )
        declaration.masse_salariale_brute = masse

        return declaration

    def _trouver_entete(self, rows: list) -> tuple:
        """Cherche la ligne d en-tete dans les 8 premieres lignes."""
        for idx in range(min(8, len(rows))):
            row = rows[idx]
            if not row:
                continue
            header = [self._normaliser_entete(c) for c in row]
            col_map = self._mapper_colonnes(header)
            has_identity = any(f in col_map for f in ("nom", "prenom", "nom_prenom", "nir", "matricule"))
            has_amount = any(f in col_map for f in ("base_brute", "net", "montant_patronal", "montant_salarial"))
            if len(col_map) >= 2 and (has_identity or has_amount):
                return idx, header
        # Fallback: first non-empty row with any mappable column
        for idx in range(min(8, len(rows))):
            row = rows[idx]
            if not row:
                continue
            header = [self._normaliser_entete(c) for c in row]
            col_map = self._mapper_colonnes(header)
            if col_map:
                return idx, header
        return None, None

    @staticmethod
    def _est_ligne_total(row_data: tuple) -> bool:
        """Detecte si une ligne est un total/sous-total a ignorer."""
        for cell in row_data:
            if cell is None:
                continue
            s = str(cell).strip().lower()
            if any(kw in s for kw in _TOTAL_KEYWORDS):
                return True
        return False

    @staticmethod
    def _normaliser_entete(cellule: Any) -> str:
        """Normalise un en-tete de colonne."""
        if cellule is None:
            return ""
        s = str(cellule).strip().lower()
        for a, b in [("\u00e9", "e"), ("\u00e8", "e"), ("\u00ea", "e"), ("\u00eb", "e"),
                      ("\u00e0", "a"), ("\u00e2", "a"), ("\u00f4", "o"), ("\u00ee", "i"),
                      ("\u00f9", "u"), ("\u00fb", "u"), ("\u00e7", "c"), ("\u00ef", "i")]:
            s = s.replace(a, b)
        s = s.replace(" ", "_").replace("-", "_").replace(".", "_")
        s = s.replace("'", "").replace("\u00b0", "")
        while "__" in s:
            s = s.replace("__", "_")
        return s.strip("_")

    @classmethod
    def _mapper_colonnes(cls, header: list[str]) -> dict[str, int]:
        """Identifie les colonnes utiles.

        Two-pass:
        1. Exact match: header == keyword
        2. Inclusion: keyword IN header (NEVER header IN keyword to avoid false matches)
        """
        mapping = {}
        used_cols = set()

        # Pass 1: Exact matches
        for i, col in enumerate(header):
            if not col or i in used_cols:
                continue
            for field_name, kws in _COLUMN_KEYWORDS.items():
                if field_name in mapping:
                    continue
                if col in kws:
                    mapping[field_name] = i
                    used_cols.add(i)
                    break

        # Pass 2: Inclusion (keyword IN header only)
        for i, col in enumerate(header):
            if not col or i in used_cols:
                continue
            for field_name, kws in _COLUMN_KEYWORDS.items():
                if field_name in mapping:
                    continue
                for kw in kws:
                    if kw in col and len(kw) >= 3:
                        mapping[field_name] = i
                        used_cols.add(i)
                        break
                if field_name in mapping:
                    break

        return mapping

    def _extraire_employe_mapped(self, mapped_row: dict, doc_id: str) -> Employe | None:
        """Extract employee from a mapped row (field_name -> value)."""
        nir = None
        matricule = None
        nom = None
        prenom = None

        # NIR
        raw_nir = mapped_row.get("nir")
        if raw_nir:
            nir = str(raw_nir).strip().replace(" ", "")
            if nir.endswith(".0"):
                nir = nir[:-2]

        # Matricule
        raw_mat = mapped_row.get("matricule")
        if raw_mat:
            matricule = str(raw_mat).strip()
            if matricule.endswith(".0"):
                matricule = matricule[:-2]

        # Nom + Prenom
        raw_nom = mapped_row.get("nom")
        if raw_nom:
            nom = str(raw_nom).strip()
        raw_prenom = mapped_row.get("prenom")
        if raw_prenom:
            prenom = str(raw_prenom).strip()

        # Combined nom_prenom
        if not nom:
            raw_combined = mapped_row.get("nom_prenom")
            if raw_combined and str(raw_combined).strip():
                combined = str(raw_combined).strip()
                parts = combined.split()
                if len(parts) >= 2:
                    upper_parts = [p for p in parts if p == p.upper() and len(p) > 1]
                    if upper_parts:
                        nom = " ".join(upper_parts)
                        prenom = " ".join(p for p in parts if p not in upper_parts)
                    else:
                        nom = parts[0]
                        prenom = " ".join(parts[1:])
                else:
                    nom = combined

        # Use matricule as fallback identifier
        if not nir and matricule:
            nir = f"MAT_{matricule}"

        if not nir and not nom:
            return None

        emp = Employe(
            nir=nir or "",
            nom=nom or "",
            prenom=prenom or "",
            source_document_id=doc_id,
        )

        raw_statut = mapped_row.get("statut")
        if raw_statut:
            s = str(raw_statut).strip().lower()
            if "cadre" in s:
                emp.statut = "cadre"
            elif "apprenti" in s or "alternant" in s:
                emp.statut = "apprenti"
            else:
                emp.statut = "non-cadre"

        raw_poste = mapped_row.get("poste")
        if raw_poste:
            emp.convention_collective = str(raw_poste).strip()

        return emp

    def _extraire_cotisation_mapped(self, mapped_row: dict, doc_id: str) -> Cotisation | None:
        """Extract cotisation from a mapped row."""
        base = mapped_row.get("base_brute")
        montant_p = mapped_row.get("montant_patronal")
        montant_s = mapped_row.get("montant_salarial")
        net = mapped_row.get("net")

        if base is None and montant_p is None and montant_s is None and net is None:
            return None

        c = Cotisation(source_document_id=doc_id)

        if base is not None:
            c.base_brute = self._to_decimal(base)
            c.assiette = c.base_brute
        if montant_p is not None:
            c.montant_patronal = self._to_decimal(montant_p)
        if montant_s is not None:
            c.montant_salarial = self._to_decimal(montant_s)

        tp = mapped_row.get("taux_patronal")
        if tp is not None:
            c.taux_patronal = self._to_decimal(tp)
            if c.taux_patronal > 1:
                c.taux_patronal = c.taux_patronal / 100

        ts = mapped_row.get("taux_salarial")
        if ts is not None:
            c.taux_salarial = self._to_decimal(ts)
            if c.taux_salarial > 1:
                c.taux_salarial = c.taux_salarial / 100

        # If only net, estimate brut
        if c.base_brute <= 0 and net is not None:
            net_val = self._to_decimal(net)
            if net_val > 0:
                c.base_brute = Decimal(str(round(float(net_val) / 0.78, 2)))
                c.assiette = c.base_brute

        return c

    @staticmethod
    def _to_decimal(val: Any) -> Decimal:
        """Convertit une valeur en Decimal."""
        if isinstance(val, Decimal):
            return val
        if isinstance(val, (int, float)):
            return Decimal(str(val))
        if isinstance(val, str):
            return parser_montant(val)
        return Decimal(str(val))
