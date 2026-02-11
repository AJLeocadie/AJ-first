"""Lecteur multi-format avec detection d'ecriture manuscrite.

Supporte :
- Images : JPEG, PNG, BMP, TIFF, GIF, WEBP
- PDF (via pdfplumber)
- Excel (via openpyxl)
- Texte brut, CSV, XML, DSN
- Documents scannes avec detection OCR

Detection manuscrit :
- Analyse des patterns irreguliers de texte
- Detection de zones a faible confiance OCR
- Avertissements utilisateur explicites
"""

import re
import base64
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Optional


class FormatFichier(str, Enum):
    PDF = "pdf"
    JPEG = "jpeg"
    PNG = "png"
    BMP = "bmp"
    TIFF = "tiff"
    GIF = "gif"
    WEBP = "webp"
    CSV = "csv"
    EXCEL = "excel"
    XML = "xml"
    DSN = "dsn"
    TEXTE = "texte"
    INCONNU = "inconnu"


# Extensions par format
EXTENSIONS_IMAGES = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".gif", ".webp"}
EXTENSIONS_PDF = {".pdf"}
EXTENSIONS_EXCEL = {".xlsx", ".xls"}
EXTENSIONS_CSV = {".csv"}
EXTENSIONS_XML = {".xml"}
EXTENSIONS_DSN = {".dsn"}
EXTENSIONS_TEXTE = {".txt", ".text", ".log"}

TOUTES_EXTENSIONS = (
    EXTENSIONS_IMAGES | EXTENSIONS_PDF | EXTENSIONS_EXCEL |
    EXTENSIONS_CSV | EXTENSIONS_XML | EXTENSIONS_DSN | EXTENSIONS_TEXTE
)


@dataclass
class AvertissementManuscrit:
    """Avertissement sur une zone manuscrite detectee."""
    zone: str
    message: str
    confiance: float  # 0-1, confiance que c'est manuscrit
    ligne_numero: int = 0


@dataclass
class ResultatLecture:
    """Resultat de la lecture d'un document."""
    texte: str = ""
    format_detecte: FormatFichier = FormatFichier.INCONNU
    nom_fichier: str = ""
    taille_octets: int = 0
    nb_pages: int = 1
    est_image: bool = False
    est_scan: bool = False
    manuscrit_detecte: bool = False
    avertissements_manuscrit: list[AvertissementManuscrit] = field(default_factory=list)
    avertissements: list[str] = field(default_factory=list)
    confiance_ocr: float = 1.0  # 1.0 = texte natif, < 0.8 = scan/manuscrit
    metadonnees: dict = field(default_factory=dict)
    donnees_structurees: list[dict] = field(default_factory=list)  # Pour Excel/CSV


# Patterns de detection manuscrit ameliores
PATTERNS_MANUSCRIT_FORT = [
    # Melange aleatoire majuscules/minuscules dans un mot
    re.compile(r'\b[a-z]+[A-Z][a-z]+[A-Z]\w*\b'),
    # Chiffres et lettres entremeles de facon inhabituelle
    re.compile(r'[A-Za-z]\d[A-Za-z]\d'),
    # Caracteres de substitution OCR frequents
    re.compile(r'[Il1|]{3,}|[O0]{3,}'),
    # Mots tronques ou incomplets
    re.compile(r'\b\w{1,2}\s\w{1,2}\s\w{1,2}\b'),
]

PATTERNS_MANUSCRIT_FAIBLE = [
    # Annotations courtes potentiellement manuscrites
    re.compile(r'^[A-Z][a-z]{0,3}\s*\d{1,6}$', re.MULTILINE),
    # Croix, coches, annotations marginales
    re.compile(r'^\s*[xXvV✓✗]\s*$', re.MULTILINE),
    # Fleches textuelles
    re.compile(r'->|-->|=>|<-'),
]

# Indicateurs de scan (vs texte natif)
INDICATEURS_SCAN = [
    # Beaucoup de caracteres speciaux de mauvais OCR
    re.compile(r'[~`\^]{2,}'),
    # Lignes de caracteres repetitifs (artefacts scan)
    re.compile(r'(.)\1{8,}'),
    # Espaces irreguliers entre mots
    re.compile(r'\w\s{3,}\w'),
]


class LecteurMultiFormat:
    """Lecteur universel de documents avec detection manuscrite."""

    def lire_fichier(self, chemin: Path) -> ResultatLecture:
        """Lit un fichier et retourne le texte + metadonnees."""
        if not chemin.exists():
            return ResultatLecture(
                avertissements=[f"Fichier introuvable : {chemin}"]
            )

        ext = chemin.suffix.lower()
        resultat = ResultatLecture(
            nom_fichier=chemin.name,
            taille_octets=chemin.stat().st_size,
        )

        # Dispatch selon le format
        if ext in EXTENSIONS_IMAGES:
            resultat = self._lire_image(chemin, resultat)
        elif ext in EXTENSIONS_PDF:
            resultat = self._lire_pdf(chemin, resultat)
        elif ext in EXTENSIONS_EXCEL:
            resultat = self._lire_excel(chemin, resultat)
        elif ext in EXTENSIONS_CSV:
            resultat = self._lire_csv(chemin, resultat)
        elif ext in EXTENSIONS_XML or ext in EXTENSIONS_DSN:
            resultat = self._lire_texte(chemin, resultat)
            resultat.format_detecte = FormatFichier.XML if ext in EXTENSIONS_XML else FormatFichier.DSN
        else:
            resultat = self._lire_texte(chemin, resultat)

        # Detection manuscrit sur le texte extrait
        if resultat.texte:
            self._detecter_manuscrit(resultat)
            self._detecter_scan(resultat)

        return resultat

    def lire_contenu_brut(self, contenu: bytes, nom_fichier: str = "") -> ResultatLecture:
        """Lit depuis un contenu en memoire (upload)."""
        ext = Path(nom_fichier).suffix.lower() if nom_fichier else ""
        resultat = ResultatLecture(
            nom_fichier=nom_fichier,
            taille_octets=len(contenu),
        )

        if ext in EXTENSIONS_IMAGES:
            resultat = self._lire_image_bytes(contenu, resultat)
        elif ext in EXTENSIONS_PDF:
            resultat = self._lire_pdf_bytes(contenu, resultat)
        elif ext in EXTENSIONS_EXCEL:
            resultat = self._lire_excel_bytes(contenu, resultat)
        elif ext in EXTENSIONS_CSV:
            texte = self._decoder_texte(contenu)
            resultat.texte = texte
            resultat.format_detecte = FormatFichier.CSV
        else:
            texte = self._decoder_texte(contenu)
            resultat.texte = texte
            resultat.format_detecte = FormatFichier.TEXTE

        if resultat.texte:
            self._detecter_manuscrit(resultat)
            self._detecter_scan(resultat)

        return resultat

    # --- Lecteurs specifiques ---

    def _lire_image(self, chemin: Path, resultat: ResultatLecture) -> ResultatLecture:
        """Lit une image et tente une extraction OCR basique."""
        ext = chemin.suffix.lower()
        resultat.est_image = True
        resultat.format_detecte = self._format_depuis_ext(ext)

        # Tenter OCR via Pillow + pytesseract si disponible
        texte_ocr = self._ocr_image_fichier(chemin)
        if texte_ocr:
            resultat.texte = texte_ocr
            resultat.confiance_ocr = 0.7
            resultat.est_scan = True
        else:
            # Pas d'OCR disponible - analyser les metadonnees image
            resultat.avertissements.append(
                "AVERTISSEMENT : Ce fichier est une image. "
                "L'extraction de texte est limitee sans moteur OCR (Tesseract). "
                "Veuillez verifier manuellement le contenu."
            )
            resultat.confiance_ocr = 0.3
            # Tenter lecture basique du texte embede dans l'image
            try:
                with open(chemin, "rb") as f:
                    data = f.read()
                resultat.texte = self._extraire_texte_image_basique(data)
            except Exception:
                resultat.texte = ""

        resultat.metadonnees["type_image"] = ext
        resultat.metadonnees["taille"] = resultat.taille_octets
        return resultat

    def _lire_image_bytes(self, contenu: bytes, resultat: ResultatLecture) -> ResultatLecture:
        """Lit une image depuis des bytes."""
        resultat.est_image = True
        ext = Path(resultat.nom_fichier).suffix.lower()
        resultat.format_detecte = self._format_depuis_ext(ext)

        texte_ocr = self._ocr_image_bytes(contenu)
        if texte_ocr:
            resultat.texte = texte_ocr
            resultat.confiance_ocr = 0.7
            resultat.est_scan = True
        else:
            resultat.avertissements.append(
                "AVERTISSEMENT : Ce fichier est une image. "
                "L'extraction de texte est limitee sans moteur OCR (Tesseract). "
                "Veuillez verifier manuellement le contenu."
            )
            resultat.confiance_ocr = 0.3
            resultat.texte = self._extraire_texte_image_basique(contenu)

        return resultat

    def _lire_pdf(self, chemin: Path, resultat: ResultatLecture) -> ResultatLecture:
        """Lit un fichier PDF."""
        resultat.format_detecte = FormatFichier.PDF
        try:
            import pdfplumber
            with pdfplumber.open(chemin) as pdf:
                pages_texte = []
                for page in pdf.pages:
                    t = page.extract_text() or ""
                    pages_texte.append(t)
                resultat.texte = "\n".join(pages_texte)
                resultat.nb_pages = len(pdf.pages)

                # Detecter si c'est un scan (peu de texte vs nb pages)
                if resultat.nb_pages > 0:
                    mots_par_page = len(resultat.texte.split()) / resultat.nb_pages
                    if mots_par_page < 20:
                        resultat.est_scan = True
                        resultat.confiance_ocr = 0.5
                        resultat.avertissements.append(
                            "AVERTISSEMENT : Ce PDF semble etre un document scanne "
                            f"(~{mots_par_page:.0f} mots/page). Le contenu textuel "
                            "peut etre incomplet ou imprecis."
                        )
        except ImportError:
            resultat.avertissements.append(
                "Module pdfplumber non disponible. Lecture PDF limitee."
            )
        except Exception as e:
            resultat.avertissements.append(f"Erreur lecture PDF : {str(e)}")

        return resultat

    def _lire_pdf_bytes(self, contenu: bytes, resultat: ResultatLecture) -> ResultatLecture:
        """Lit un PDF depuis des bytes."""
        resultat.format_detecte = FormatFichier.PDF
        try:
            import pdfplumber
            import io
            with pdfplumber.open(io.BytesIO(contenu)) as pdf:
                pages_texte = []
                for page in pdf.pages:
                    t = page.extract_text() or ""
                    pages_texte.append(t)
                resultat.texte = "\n".join(pages_texte)
                resultat.nb_pages = len(pdf.pages)

                if resultat.nb_pages > 0:
                    mots_par_page = len(resultat.texte.split()) / resultat.nb_pages
                    if mots_par_page < 20:
                        resultat.est_scan = True
                        resultat.confiance_ocr = 0.5
                        resultat.avertissements.append(
                            "AVERTISSEMENT : Ce PDF semble etre un document scanne."
                        )
        except ImportError:
            resultat.avertissements.append(
                "Module pdfplumber non disponible."
            )
        except Exception as e:
            resultat.avertissements.append(f"Erreur lecture PDF : {str(e)}")

        return resultat

    def _lire_excel(self, chemin: Path, resultat: ResultatLecture) -> ResultatLecture:
        """Lit un fichier Excel."""
        resultat.format_detecte = FormatFichier.EXCEL
        try:
            import openpyxl
            wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
            lignes_texte = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    vals = [str(c) if c is not None else "" for c in row]
                    ligne = " | ".join(v for v in vals if v)
                    if ligne.strip():
                        lignes_texte.append(ligne)
                        resultat.donnees_structurees.append(
                            {f"col_{i}": str(c) if c is not None else "" for i, c in enumerate(row)}
                        )
            wb.close()
            resultat.texte = "\n".join(lignes_texte)
            resultat.nb_pages = len(wb.sheetnames)
        except ImportError:
            resultat.avertissements.append(
                "Module openpyxl non disponible. Lecture Excel impossible."
            )
        except Exception as e:
            resultat.avertissements.append(f"Erreur lecture Excel : {str(e)}")

        return resultat

    def _lire_excel_bytes(self, contenu: bytes, resultat: ResultatLecture) -> ResultatLecture:
        """Lit un Excel depuis des bytes."""
        resultat.format_detecte = FormatFichier.EXCEL
        try:
            import openpyxl
            import io
            wb = openpyxl.load_workbook(io.BytesIO(contenu), read_only=True, data_only=True)
            lignes_texte = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    vals = [str(c) if c is not None else "" for c in row]
                    ligne = " | ".join(v for v in vals if v)
                    if ligne.strip():
                        lignes_texte.append(ligne)
                        resultat.donnees_structurees.append(
                            {f"col_{i}": str(c) if c is not None else "" for i, c in enumerate(row)}
                        )
            wb.close()
            resultat.texte = "\n".join(lignes_texte)
        except ImportError:
            resultat.avertissements.append(
                "Module openpyxl non disponible."
            )
        except Exception as e:
            resultat.avertissements.append(f"Erreur lecture Excel : {str(e)}")

        return resultat

    def _lire_csv(self, chemin: Path, resultat: ResultatLecture) -> ResultatLecture:
        """Lit un fichier CSV."""
        resultat.format_detecte = FormatFichier.CSV
        try:
            try:
                with open(chemin, "r", encoding="utf-8-sig") as f:
                    contenu = f.read()
            except UnicodeDecodeError:
                with open(chemin, "r", encoding="latin-1") as f:
                    contenu = f.read()
            resultat.texte = contenu

            # Parser en structure
            import csv
            import io
            reader = csv.DictReader(io.StringIO(contenu))
            for row in reader:
                resultat.donnees_structurees.append(dict(row))

        except Exception as e:
            resultat.avertissements.append(f"Erreur lecture CSV : {str(e)}")

        return resultat

    def _lire_texte(self, chemin: Path, resultat: ResultatLecture) -> ResultatLecture:
        """Lit un fichier texte brut."""
        resultat.format_detecte = FormatFichier.TEXTE
        try:
            try:
                resultat.texte = chemin.read_text(encoding="utf-8")
            except UnicodeDecodeError:
                resultat.texte = chemin.read_text(encoding="latin-1")
        except Exception as e:
            resultat.avertissements.append(f"Erreur lecture fichier : {str(e)}")

        return resultat

    # --- Detection manuscrit ---

    def _detecter_manuscrit(self, resultat: ResultatLecture):
        """Detecte les zones manuscrites dans le texte extrait."""
        if not resultat.texte:
            return

        for i, ligne in enumerate(resultat.texte.split("\n"), 1):
            ligne = ligne.strip()
            if not ligne:
                continue

            # Patterns forts de manuscrit
            for pattern in PATTERNS_MANUSCRIT_FORT:
                if pattern.search(ligne):
                    resultat.avertissements_manuscrit.append(
                        AvertissementManuscrit(
                            zone=ligne[:100],
                            message="Ecriture manuscrite detectee (confiance elevee)",
                            confiance=0.85,
                            ligne_numero=i,
                        )
                    )
                    break

            # Analyse heuristique : ratio majuscules irregulier
            upper = sum(1 for c in ligne if c.isupper())
            lower = sum(1 for c in ligne if c.islower())
            total = upper + lower
            if total > 8:
                ratio = upper / total
                if 0.25 < ratio < 0.65:
                    # Verifier que ce n'est pas un en-tete normal
                    if not any(kw in ligne.upper() for kw in [
                        "FACTURE", "TOTAL", "TVA", "SIRET", "N°", "DATE",
                        "MONTANT", "SARL", "SAS", "SA ", "EURL",
                    ]):
                        resultat.avertissements_manuscrit.append(
                            AvertissementManuscrit(
                                zone=ligne[:100],
                                message="Zone potentiellement manuscrite (casse irreguliere)",
                                confiance=0.55,
                                ligne_numero=i,
                            )
                        )

            # Patterns faibles
            for pattern in PATTERNS_MANUSCRIT_FAIBLE:
                if pattern.search(ligne):
                    # Verifier pas deja detecte
                    if not any(a.ligne_numero == i for a in resultat.avertissements_manuscrit):
                        resultat.avertissements_manuscrit.append(
                            AvertissementManuscrit(
                                zone=ligne[:100],
                                message="Annotation manuscrite possible",
                                confiance=0.4,
                                ligne_numero=i,
                            )
                        )
                    break

        # Determiner si le document contient du manuscrit significatif
        if resultat.avertissements_manuscrit:
            forte_confiance = [a for a in resultat.avertissements_manuscrit if a.confiance >= 0.7]
            if forte_confiance:
                resultat.manuscrit_detecte = True
                resultat.avertissements.append(
                    f"AVERTISSEMENT MANUSCRIT : {len(forte_confiance)} zone(s) "
                    "d'ecriture manuscrite detectee(s) avec forte confiance. "
                    "Les donnees extraites de ces zones peuvent etre imprecises. "
                    "Verification manuelle recommandee."
                )
            elif len(resultat.avertissements_manuscrit) >= 3:
                resultat.manuscrit_detecte = True
                resultat.avertissements.append(
                    f"AVERTISSEMENT MANUSCRIT : {len(resultat.avertissements_manuscrit)} zone(s) "
                    "potentiellement manuscrites detectees. "
                    "Verification manuelle recommandee."
                )

    def _detecter_scan(self, resultat: ResultatLecture):
        """Detecte si le document est un scan."""
        if not resultat.texte or resultat.est_image:
            return

        score_scan = 0
        for pattern in INDICATEURS_SCAN:
            if pattern.search(resultat.texte):
                score_scan += 1

        # Ratio de caracteres speciaux
        total = len(resultat.texte)
        if total > 50:
            speciaux = sum(1 for c in resultat.texte if not c.isalnum() and not c.isspace() and c not in ".,;:!?'-/()")
            ratio_speciaux = speciaux / total
            if ratio_speciaux > 0.15:
                score_scan += 2

        if score_scan >= 2 and not resultat.est_scan:
            resultat.est_scan = True
            resultat.confiance_ocr = min(resultat.confiance_ocr, 0.6)
            resultat.avertissements.append(
                "Ce document semble avoir ete scanne. "
                "La qualite de l'extraction depend de la qualite du scan."
            )

    # --- Utilitaires ---

    def _ocr_image_fichier(self, chemin: Path) -> str:
        """Tente OCR sur un fichier image via pytesseract."""
        try:
            from PIL import Image
            import pytesseract
            img = Image.open(chemin)
            return pytesseract.image_to_string(img, lang="fra")
        except ImportError:
            return ""
        except Exception:
            return ""

    def _ocr_image_bytes(self, contenu: bytes) -> str:
        """Tente OCR sur des bytes d'image."""
        try:
            from PIL import Image
            import pytesseract
            import io
            img = Image.open(io.BytesIO(contenu))
            return pytesseract.image_to_string(img, lang="fra")
        except ImportError:
            return ""
        except Exception:
            return ""

    def _extraire_texte_image_basique(self, data: bytes) -> str:
        """Extraction basique de texte depuis une image (sans OCR complet).

        Recherche les chaines ASCII/UTF-8 dans les donnees brutes.
        """
        texte_fragments = []
        # Rechercher les sequences de caracteres imprimables
        current = []
        for byte in data:
            if 32 <= byte < 127:
                current.append(chr(byte))
            elif current:
                mot = "".join(current)
                if len(mot) >= 4:  # Filtrer le bruit
                    texte_fragments.append(mot)
                current = []

        if current:
            mot = "".join(current)
            if len(mot) >= 4:
                texte_fragments.append(mot)

        # Filtrer les patterns pertinents (mots-cles comptables/fiscaux)
        mots_cles = {
            "facture", "invoice", "total", "montant", "date", "siret",
            "tva", "ttc", "avoir", "numero", "client", "fournisseur",
            "paiement", "reglement", "echeance", "reference",
            "brut", "net", "cotisation", "salaire", "paie",
        }
        pertinents = []
        for frag in texte_fragments:
            if any(mc in frag.lower() for mc in mots_cles):
                pertinents.append(frag)

        return " ".join(pertinents[:200])

    def _decoder_texte(self, contenu: bytes) -> str:
        """Decode des bytes en texte avec fallback."""
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                return contenu.decode(enc)
            except (UnicodeDecodeError, UnicodeError):
                continue
        return contenu.decode("latin-1", errors="replace")

    def _format_depuis_ext(self, ext: str) -> FormatFichier:
        """Determine le format depuis l'extension."""
        ext = ext.lower()
        mapping = {
            ".jpg": FormatFichier.JPEG, ".jpeg": FormatFichier.JPEG,
            ".png": FormatFichier.PNG, ".bmp": FormatFichier.BMP,
            ".tiff": FormatFichier.TIFF, ".tif": FormatFichier.TIFF,
            ".gif": FormatFichier.GIF, ".webp": FormatFichier.WEBP,
            ".pdf": FormatFichier.PDF,
            ".xlsx": FormatFichier.EXCEL, ".xls": FormatFichier.EXCEL,
            ".csv": FormatFichier.CSV,
            ".xml": FormatFichier.XML, ".dsn": FormatFichier.DSN,
            ".txt": FormatFichier.TEXTE,
        }
        return mapping.get(ext, FormatFichier.INCONNU)
