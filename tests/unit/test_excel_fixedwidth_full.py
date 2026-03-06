"""Comprehensive tests for excel_parser.py and fixedwidth_parser.py covering all uncovered lines."""

import tempfile
from datetime import date
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch, MagicMock

import openpyxl
import pytest

from urssaf_analyzer.models.documents import Document, FileType
from urssaf_analyzer.parsers.excel_parser import ExcelParser, _COLUMN_KEYWORDS
from urssaf_analyzer.parsers.fixedwidth_parser import (
    FixedWidthParser,
    _parse_sage_date,
    _parse_ciel_date,
    _parse_fixed_montant,
    _SAGE_PNM_LINE_LEN,
    _CIEL_LINE_LEN_MIN,
)
from urssaf_analyzer.core.exceptions import ParseError


# ============================================================
# Helpers
# ============================================================

def _make_doc(name="test.xlsx", ftype=FileType.EXCEL, path=None):
    return Document(
        nom_fichier=name,
        chemin=path or Path("/tmp/fake"),
        type_fichier=ftype,
        hash_sha256="a" * 64,
        taille_octets=1024,
    )


def _write_excel(path, sheets):
    """Write an Excel workbook with the given sheets.

    sheets: dict of {sheet_name: list_of_rows} where each row is a list/tuple.
    """
    wb = openpyxl.Workbook()
    first = True
    for name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        for row in rows:
            ws.append(list(row))
    wb.save(path)
    wb.close()


def _sage_pnm_line(journal="OD ", date_str="150124", type_l="OD",
                    compte="4210000000000", flag=" ", tiers="             ",
                    piece="BP0001       ", libelle="Salaire brut janvier  ",
                    mode=" ", echeance="310124", debit="    3500.00  ",
                    credit="       0.00  "):
    """Build a SAGE PNM fixed-width line (109 chars)."""
    line = (
        f"{journal:<3}"
        f"{date_str:<6}"
        f"{type_l:<2}"
        f"{compte:<13}"
        f"{flag:<1}"
        f"{tiers:<13}"
        f"{piece:<13}"
        f"{libelle:<25}"
        f"{mode:<1}"
        f"{echeance:<6}"
        f"{debit:>13}"
        f"{credit:>13}"
    )
    return line


def _ciel_ximport_line(n_mouvement="00001", journal="OD", date_str="20240115",
                       echeance="20240131", piece="BP0001      ",
                       compte="42100000000", libelle="Salaire brut janvier  ",
                       montant="     3500.00 ", sens="D",
                       pointage="            ", analytique="      ",
                       libelle_compte="Remuneration du personnel         ",
                       euro="E"):
    """Build a CIEL XIMPORT fixed-width line (~138 chars)."""
    line = (
        f"{n_mouvement:<5}"
        f"{journal:<2}"
        f"{date_str:<8}"
        f"{echeance:<8}"
        f"{piece:<12}"
        f"{compte:<11}"
        f"{libelle:<25}"
        f"{montant:>13}"
        f"{sens:<1}"
        f"{pointage:<12}"
        f"{analytique:<6}"
        f"{libelle_compte:<34}"
        f"{euro:<1}"
    )
    return line


# ============================================================
# ExcelParser Tests
# ============================================================

class TestExcelParserPeutTraiter:
    def test_xlsx(self, tmp_path):
        p = ExcelParser()
        assert p.peut_traiter(Path("test.xlsx")) is True

    def test_xls(self):
        p = ExcelParser()
        assert p.peut_traiter(Path("test.xls")) is True

    def test_csv_rejected(self):
        p = ExcelParser()
        assert p.peut_traiter(Path("test.csv")) is False


class TestExcelParserExtraireMetadata:
    """Covers lines 229-238 (metadata extraction)."""

    def test_extraire_metadata_returns_sheets(self, tmp_path):
        """Covers lines 229-238 using a mock workbook with dimensions attribute."""
        path = tmp_path / "meta.xlsx"
        _write_excel(path, {
            "Paie": [["Nom", "Brut"], ["Dupont", 3000]],
            "Recap": [["Total"], [6000]],
        })
        # Mock openpyxl.load_workbook to return a wb with .dimensions on sheets
        mock_ws1 = MagicMock()
        mock_ws1.dimensions = "A1:B2"
        mock_ws2 = MagicMock()
        mock_ws2.dimensions = "A1:A2"
        mock_wb = MagicMock()
        mock_wb.sheetnames = ["Paie", "Recap"]
        mock_wb.__getitem__ = lambda self, k: mock_ws1 if k == "Paie" else mock_ws2

        p = ExcelParser()
        with patch("urssaf_analyzer.parsers.excel_parser.openpyxl") as mock_openpyxl:
            mock_openpyxl.load_workbook.return_value = mock_wb
            meta = p.extraire_metadata(path)
        assert meta["format"] == "excel"
        assert meta["feuilles"] == ["Paie", "Recap"]
        assert meta["nb_feuilles"] == 2
        assert meta["feuille_Paie_dims"] == "A1:B2"
        assert meta["feuille_Recap_dims"] == "A1:A2"

    def test_extraire_metadata_no_openpyxl(self, tmp_path):
        """Covers lines 38-39 (HAS_OPENPYXL = False branch)."""
        path = tmp_path / "meta.xlsx"
        _write_excel(path, {"S1": [["a"]]})
        p = ExcelParser()
        with patch("urssaf_analyzer.parsers.excel_parser.HAS_OPENPYXL", False):
            meta = p.extraire_metadata(path)
        assert meta == {"erreur": "openpyxl non installe"}

    def test_extraire_metadata_corrupt_file(self, tmp_path):
        path = tmp_path / "bad.xlsx"
        path.write_bytes(b"not an excel file")
        p = ExcelParser()
        meta = p.extraire_metadata(path)
        assert "erreur" in meta


class TestExcelParserParserFlow:
    """Covers lines 246-259 (full parser flow)."""

    def test_parser_no_openpyxl_raises(self, tmp_path):
        path = tmp_path / "test.xlsx"
        _write_excel(path, {"S": [["a"]]})
        p = ExcelParser()
        doc = _make_doc(path=path)
        with patch("urssaf_analyzer.parsers.excel_parser.HAS_OPENPYXL", False):
            with pytest.raises(ParseError, match="openpyxl"):
                p.parser(path, doc)

    def test_parser_corrupt_file_raises(self, tmp_path):
        path = tmp_path / "corrupt.xlsx"
        path.write_bytes(b"corrupted data")
        p = ExcelParser()
        doc = _make_doc(path=path)
        with pytest.raises(ParseError, match="Impossible de lire"):
            p.parser(path, doc)

    def test_parser_full_flow_single_employee(self, tmp_path):
        """Full parser flow with one employee - covers _parser_feuille, _trouver_entete,
        _mapper_colonnes, _extraire_employe_mapped, _extraire_cotisation_mapped."""
        path = tmp_path / "bulletin.xlsx"
        _write_excel(path, {
            "Bulletin": [
                ("Nom", "Prenom", "Salaire Brut", "Net", "Montant Patronal", "Montant Salarial"),
                ("DUPONT", "Jean", 3500.00, 2730.00, 525.00, 245.00),
            ]
        })
        p = ExcelParser()
        doc = _make_doc(path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        decl = decls[0]
        assert decl.type_declaration == "bulletin"
        assert len(decl.employes) == 1
        assert decl.employes[0].nom == "DUPONT"
        assert decl.employes[0].prenom == "Jean"
        assert len(decl.cotisations) >= 1
        assert decl.cotisations[0].base_brute == Decimal("3500.0")

    def test_parser_full_flow_multiple_employees(self, tmp_path):
        """Livre de paie flow with multiple employees - covers masse salariale calc."""
        path = tmp_path / "livre.xlsx"
        _write_excel(path, {
            "Paie": [
                ("Nom", "Prenom", "NIR", "Salaire Brut", "Net", "Montant Patronal"),
                ("DUPONT", "Jean", "1850175000001 23", 3500.00, 2730.00, 525.00),
                ("MARTIN", "Sophie", "2890175000002 45", 4000.00, 3120.00, 600.00),
            ]
        })
        p = ExcelParser()
        doc = _make_doc(path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        decl = decls[0]
        assert decl.type_declaration == "livre_de_paie"
        assert len(decl.employes) == 2
        assert decl.effectif_declare == 2
        assert decl.metadata["type_document"] == "livre_de_paie"

    def test_parser_empty_sheet_skipped(self, tmp_path):
        """Sheet with only one row (header only) returns no declarations."""
        path = tmp_path / "empty.xlsx"
        _write_excel(path, {
            "S1": [("Nom", "Brut")],  # Only header, no data
        })
        p = ExcelParser()
        doc = _make_doc(path=path)
        decls = p.parser(path, doc)
        assert decls == []

    def test_parser_no_mappable_headers(self, tmp_path):
        """Sheet with unrecognizable headers returns no declarations."""
        path = tmp_path / "nomatch.xlsx"
        _write_excel(path, {
            "S1": [
                ("XYZ", "ABC", "DEF"),
                (1, 2, 3),
            ]
        })
        p = ExcelParser()
        doc = _make_doc(path=path)
        decls = p.parser(path, doc)
        assert decls == []


class TestParserFeuille:
    """Covers lines 273-375 (_parser_feuille internals)."""

    def test_empty_rows_skipped(self, tmp_path):
        """Rows that are all None are skipped."""
        path = tmp_path / "gaps.xlsx"
        _write_excel(path, {
            "S1": [
                ("Nom", "Prenom", "Salaire Brut"),
                (None, None, None),
                ("DUPONT", "Jean", 3500),
            ]
        })
        p = ExcelParser()
        doc = _make_doc(path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        assert len(decls[0].employes) == 1

    def test_total_line_ignored(self, tmp_path):
        """Lines containing total keywords are skipped - covers line 405 (None cell)."""
        path = tmp_path / "totals.xlsx"
        _write_excel(path, {
            "S1": [
                ("Nom", "Prenom", "Salaire Brut"),
                ("DUPONT", "Jean", 3500),
                ("TOTAL", None, 3500),
            ]
        })
        p = ExcelParser()
        doc = _make_doc(path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        assert len(decls[0].employes) == 1

    def test_net_per_employee_stored(self, tmp_path):
        """Net values are stored per employee."""
        path = tmp_path / "net.xlsx"
        _write_excel(path, {
            "S1": [
                ("Nom", "Prenom", "Salaire Brut", "Net"),
                ("DUPONT", "Jean", 3500, 2730),
            ]
        })
        p = ExcelParser()
        doc = _make_doc(path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        net_map = decls[0].metadata.get("net_par_employe", {})
        assert len(net_map) == 1

    def test_masse_salariale_no_employes_with_cotisations(self, tmp_path):
        """Covers lines 352-357: cotisations exist but no employees detected."""
        path = tmp_path / "noname.xlsx"
        _write_excel(path, {
            "S1": [
                ("Salaire Brut", "Montant Patronal"),
                (3500, 525),
                (4000, 600),
            ]
        })
        p = ExcelParser()
        doc = _make_doc(path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        # masse computed from unique bruts
        assert decls[0].masse_salariale_brute > 0

    def test_employes_meta_stored(self, tmp_path):
        """Employee metadata (heures, poste, statut, service) is stored."""
        path = tmp_path / "meta_emp.xlsx"
        _write_excel(path, {
            "S1": [
                ("Nom", "Prenom", "Salaire Brut", "Heures", "Poste", "Statut", "Service"),
                ("DUPONT", "Jean", 3500, 151.67, "Comptable", "Cadre", "Finance"),
            ]
        })
        p = ExcelParser()
        doc = _make_doc(path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        meta = decls[0].metadata.get("employes_meta", {})
        assert len(meta) == 1
        emp_meta = list(meta.values())[0]
        assert "heures" in emp_meta
        assert "poste" in emp_meta
        assert "statut" in emp_meta
        assert "service" in emp_meta


class TestTrouverEntete:
    """Covers lines 382, 393, 397 (_trouver_entete)."""

    def test_header_in_first_row(self):
        p = ExcelParser()
        rows = [
            ("Nom", "Prenom", "Salaire Brut"),
            ("DUPONT", "Jean", 3500),
        ]
        idx, header = p._trouver_entete(rows)
        assert idx == 0

    def test_header_in_later_row(self):
        p = ExcelParser()
        rows = [
            ("Societe XYZ", None, None),
            ("Bulletin de paie", None, None),
            ("Nom", "Prenom", "Salaire Brut"),
            ("DUPONT", "Jean", 3500),
        ]
        idx, header = p._trouver_entete(rows)
        assert idx == 2

    def test_fallback_to_any_mappable(self):
        """Covers line 397: fallback when no identity/amount but some column maps."""
        p = ExcelParser()
        rows = [
            (None, None),
            ("Heures", "Service"),
            (151, "Finance"),
        ]
        idx, header = p._trouver_entete(rows)
        assert idx == 1

    def test_no_header_found(self):
        p = ExcelParser()
        rows = [
            ("XYZ", "ABC"),
            (1, 2),
        ]
        idx, header = p._trouver_entete(rows)
        assert idx is None
        assert header is None

    def test_empty_row_skipped_in_header_search(self):
        """Covers lines 382, 393: None/empty rows skipped."""
        p = ExcelParser()
        rows = [
            None,
            (),
            ("Nom", "Salaire Brut"),
            ("DUPONT", 3500),
        ]
        idx, header = p._trouver_entete(rows)
        assert idx == 2


class TestMapperColonnes:
    """Covers _mapper_colonnes with exact and inclusion matching."""

    def test_exact_match(self):
        header = ["nom", "prenom", "brut"]
        mapping = ExcelParser._mapper_colonnes(header)
        assert "nom" in mapping
        assert "prenom" in mapping
        assert "base_brute" in mapping

    def test_inclusion_match(self):
        """keyword IN header (not reverse)."""
        header = ["nom_du_salarie_complet", "montant_brut_total"]
        mapping = ExcelParser._mapper_colonnes(header)
        assert "nom" in mapping
        assert "base_brute" in mapping

    def test_empty_header_skipped(self):
        header = ["", "nom", "", "brut"]
        mapping = ExcelParser._mapper_colonnes(header)
        assert "nom" in mapping
        assert "base_brute" in mapping

    def test_no_match(self):
        header = ["xx", "yy"]
        mapping = ExcelParser._mapper_colonnes(header)
        assert mapping == {}


class TestNormaliserEntete:
    """Covers line 424 (double underscore normalization)."""

    def test_accents_removed(self):
        assert ExcelParser._normaliser_entete("Salaire Brut") == "salaire_brut"

    def test_special_chars(self):
        assert ExcelParser._normaliser_entete("N\u00b0 SS") == "n_ss"

    def test_double_underscore(self):
        """Covers line 424: while '__' loop."""
        result = ExcelParser._normaliser_entete("Nom - - Prenom")
        assert "__" not in result

    def test_none_value(self):
        assert ExcelParser._normaliser_entete(None) == ""

    def test_accented_chars(self):
        result = ExcelParser._normaliser_entete("\u00e9ch\u00e9ance")
        assert result == "echeance"


class TestExtraireEmployeMapped:
    """Covers lines 479, 490, 512-515, 539, 543 (_extraire_employe_mapped)."""

    def setup_method(self):
        self.parser = ExcelParser()
        self.doc_id = "test-doc-id"

    def test_nom_prenom_separate(self):
        mapped = {"nom": "DUPONT", "prenom": "Jean"}
        emp = self.parser._extraire_employe_mapped(mapped, self.doc_id)
        assert emp is not None
        assert emp.nom == "DUPONT"
        assert emp.prenom == "Jean"

    def test_nom_prenom_combined_with_uppercase_detection(self):
        """Covers lines 507-510: combined name with uppercase parts."""
        mapped = {"nom_prenom": "DUPONT Jean-Pierre"}
        emp = self.parser._extraire_employe_mapped(mapped, self.doc_id)
        assert emp is not None
        assert "DUPONT" in emp.nom

    def test_nom_prenom_combined_no_uppercase(self):
        """Covers lines 512-513: no uppercase parts in combined name."""
        mapped = {"nom_prenom": "dupont jean"}
        emp = self.parser._extraire_employe_mapped(mapped, self.doc_id)
        assert emp is not None
        assert emp.nom == "dupont"
        assert emp.prenom == "jean"

    def test_nom_prenom_combined_single_word(self):
        """Covers lines 514-515: single word combined name."""
        mapped = {"nom_prenom": "Durand"}
        emp = self.parser._extraire_employe_mapped(mapped, self.doc_id)
        assert emp is not None
        assert emp.nom == "Durand"

    def test_nir_with_float_suffix(self):
        """Covers line 479: NIR ending with .0"""
        mapped = {"nir": "1850175000001.0", "nom": "DUPONT"}
        emp = self.parser._extraire_employe_mapped(mapped, self.doc_id)
        assert emp is not None
        assert not emp.nir.endswith(".0")

    def test_matricule_as_fallback(self):
        """Covers line 490, 518-519: matricule with .0 suffix, used as fallback."""
        mapped = {"matricule": "12345.0", "nom": "DUPONT"}
        emp = self.parser._extraire_employe_mapped(mapped, self.doc_id)
        assert emp is not None
        assert emp.nir == "MAT_12345"

    def test_statut_cadre(self):
        """Covers line 535."""
        mapped = {"nom": "DUPONT", "statut": "Cadre superieur"}
        emp = self.parser._extraire_employe_mapped(mapped, self.doc_id)
        assert emp.statut == "cadre"

    def test_statut_apprenti(self):
        """Covers line 537."""
        mapped = {"nom": "DUPONT", "statut": "Apprenti BTS"}
        emp = self.parser._extraire_employe_mapped(mapped, self.doc_id)
        assert emp.statut == "apprenti"

    def test_statut_alternant(self):
        mapped = {"nom": "DUPONT", "statut": "Alternant"}
        emp = self.parser._extraire_employe_mapped(mapped, self.doc_id)
        assert emp.statut == "apprenti"

    def test_statut_non_cadre(self):
        """Covers line 539: default statut."""
        mapped = {"nom": "DUPONT", "statut": "Employe"}
        emp = self.parser._extraire_employe_mapped(mapped, self.doc_id)
        assert emp.statut == "non-cadre"

    def test_poste_stored(self):
        """Covers line 543: poste sets convention_collective."""
        mapped = {"nom": "DUPONT", "poste": "Comptable"}
        emp = self.parser._extraire_employe_mapped(mapped, self.doc_id)
        assert emp.convention_collective == "Comptable"

    def test_no_identity_returns_none(self):
        mapped = {"base_brute": 3500}
        emp = self.parser._extraire_employe_mapped(mapped, self.doc_id)
        assert emp is None

    def test_matricule_only_no_nom(self):
        """Matricule only with no nom - MAT_ prefix used for nir."""
        mapped = {"matricule": "99"}
        emp = self.parser._extraire_employe_mapped(mapped, self.doc_id)
        assert emp is not None
        assert emp.nir == "MAT_99"


class TestExtraireCotisationMapped:
    """Covers lines 575-577, 597 (_extraire_cotisation_mapped)."""

    def setup_method(self):
        self.parser = ExcelParser()
        self.doc_id = "test-doc-id"

    def test_basic_cotisation(self):
        mapped = {"base_brute": 3500, "montant_patronal": 525, "montant_salarial": 245}
        cot = self.parser._extraire_cotisation_mapped(mapped, self.doc_id)
        assert cot is not None
        assert cot.base_brute == Decimal("3500")
        assert cot.montant_patronal == Decimal("525")
        assert cot.montant_salarial == Decimal("245")

    def test_no_amounts_returns_none(self):
        mapped = {"nom": "DUPONT"}
        cot = self.parser._extraire_cotisation_mapped(mapped, self.doc_id)
        assert cot is None

    def test_taux_patronal_normalization_percent(self):
        """Covers lines 570-571: taux > 1 gets divided by 100."""
        mapped = {"base_brute": 3500, "taux_patronal": 15.0}
        cot = self.parser._extraire_cotisation_mapped(mapped, self.doc_id)
        assert cot.taux_patronal == Decimal("15.0") / 100

    def test_taux_patronal_already_decimal(self):
        mapped = {"base_brute": 3500, "taux_patronal": 0.15}
        cot = self.parser._extraire_cotisation_mapped(mapped, self.doc_id)
        assert cot.taux_patronal == Decimal("0.15")

    def test_taux_salarial_normalization(self):
        """Covers lines 575-577: taux_salarial > 1 gets divided by 100."""
        mapped = {"base_brute": 3500, "taux_salarial": 7.5}
        cot = self.parser._extraire_cotisation_mapped(mapped, self.doc_id)
        assert cot.taux_salarial == Decimal("7.5") / 100

    def test_taux_salarial_already_fraction(self):
        mapped = {"base_brute": 3500, "taux_salarial": 0.075}
        cot = self.parser._extraire_cotisation_mapped(mapped, self.doc_id)
        assert cot.taux_salarial == Decimal("0.075")

    def test_net_only_estimation(self):
        """Covers lines 580-584: only net provided, brut estimated."""
        mapped = {"net": 2730}
        cot = self.parser._extraire_cotisation_mapped(mapped, self.doc_id)
        assert cot is not None
        assert cot.base_brute > 0
        assert cot.assiette == cot.base_brute
        expected = Decimal(str(round(2730 / 0.78, 2)))
        assert cot.base_brute == expected


class TestToDecimal:
    """Covers line 597 (_to_decimal with different types)."""

    def test_decimal_passthrough(self):
        assert ExcelParser._to_decimal(Decimal("42.5")) == Decimal("42.5")

    def test_int_value(self):
        assert ExcelParser._to_decimal(42) == Decimal("42")

    def test_float_value(self):
        assert ExcelParser._to_decimal(3.14) == Decimal("3.14")

    def test_string_value(self):
        result = ExcelParser._to_decimal("1 234,56")
        assert result > 0

    def test_other_type(self):
        """Covers line 597: fallback for unknown type via str()."""
        # Use a custom object that has __str__
        class FakeNum:
            def __str__(self):
                return "99"
        result = ExcelParser._to_decimal(FakeNum())
        assert result == Decimal("99")


class TestEstLigneTotal:
    def test_total_detected(self):
        assert ExcelParser._est_ligne_total(("TOTAL", 3500)) is True

    def test_sous_total_detected(self):
        assert ExcelParser._est_ligne_total(("Sous-Total", 1000)) is True

    def test_cumul_detected(self):
        assert ExcelParser._est_ligne_total(("Cumuls annuels", None, 42000)) is True

    def test_normal_row(self):
        assert ExcelParser._est_ligne_total(("DUPONT", "Jean", 3500)) is False

    def test_none_cells_skipped(self):
        """Covers line 405: None cells in row."""
        assert ExcelParser._est_ligne_total((None, None, "TOTAL")) is True
        assert ExcelParser._est_ligne_total((None, None, None)) is False


# ============================================================
# FixedWidthParser Tests
# ============================================================

class TestParseSageDate:
    """Covers lines 81-82 (_parse_sage_date edge cases)."""

    def test_valid_date(self):
        assert _parse_sage_date("150124") == date(2024, 1, 15)

    def test_year_after_50(self):
        """Year >= 50 maps to 1900s."""
        assert _parse_sage_date("150175") == date(1975, 1, 15)

    def test_year_before_50(self):
        """Year < 50 maps to 2000s."""
        assert _parse_sage_date("150125") == date(2025, 1, 15)

    def test_wrong_length(self):
        assert _parse_sage_date("1501") is None

    def test_non_digit(self):
        assert _parse_sage_date("15AB24") is None

    def test_invalid_date_values(self):
        """Covers lines 81-82: ValueError from invalid month/day."""
        assert _parse_sage_date("321324") is None  # day 32, month 13

    def test_with_spaces_valid(self):
        # Spaces are stripped, then "150124" is valid
        assert _parse_sage_date("  150124  ") == date(2024, 1, 15)

    def test_blank_string(self):
        assert _parse_sage_date("      ") is None


class TestParseCielDate:
    """Covers lines 92-93 (_parse_ciel_date edge cases)."""

    def test_valid_date(self):
        assert _parse_ciel_date("20240115") == date(2024, 1, 15)

    def test_wrong_length(self):
        assert _parse_ciel_date("2024011") is None

    def test_non_digit(self):
        assert _parse_ciel_date("2024ABCD") is None

    def test_invalid_date_values(self):
        """Covers lines 92-93: ValueError from bad date."""
        assert _parse_ciel_date("20241332") is None  # month 13

    def test_empty(self):
        assert _parse_ciel_date("") is None

    def test_spaces(self):
        assert _parse_ciel_date("        ") is None


class TestParseFixedMontant:
    """Covers _parse_fixed_montant."""

    def test_valid_amount(self):
        assert _parse_fixed_montant("  3500.00  ") == Decimal("3500.00")

    def test_comma_decimal(self):
        assert _parse_fixed_montant("3500,00") == Decimal("3500.00")

    def test_empty_string(self):
        assert _parse_fixed_montant("") == Decimal("0")

    def test_spaces_only(self):
        assert _parse_fixed_montant("     ") == Decimal("0")

    def test_invalid_returns_zero(self):
        assert _parse_fixed_montant("abc") == Decimal("0")


class TestFixedWidthParserPeutTraiter:
    """Covers lines 135-136, 151 (peut_traiter for .txt files)."""

    def test_pnm_file(self):
        p = FixedWidthParser()
        assert p.peut_traiter(Path("export.pnm")) is True

    def test_non_txt_non_pnm_rejected(self):
        p = FixedWidthParser()
        assert p.peut_traiter(Path("export.csv")) is False

    def test_txt_with_sage_content(self, tmp_path):
        """Covers lines 118-122: .txt file with SAGE PNM content."""
        path = tmp_path / "export.txt"
        lines = [_sage_pnm_line() + "\n" for _ in range(5)]
        path.write_text("".join(lines), encoding="cp1252")
        p = FixedWidthParser()
        assert p.peut_traiter(path) is True

    def test_txt_with_non_fixed_content(self, tmp_path):
        path = tmp_path / "plain.txt"
        path.write_text("This is just a plain text file.\nNothing special.\n")
        p = FixedWidthParser()
        assert p.peut_traiter(path) is False

    def test_txt_unreadable(self, tmp_path):
        """Covers line 122-123: exception reading file."""
        p = FixedWidthParser()
        assert p.peut_traiter(Path("/nonexistent/file.txt")) is False


class TestFixedWidthParserExtraireMetadata:
    """Covers lines 135-136."""

    def test_sage_metadata(self, tmp_path):
        path = tmp_path / "sage.pnm"
        lines = [_sage_pnm_line() + "\n" for _ in range(5)]
        path.write_text("".join(lines), encoding="cp1252")
        p = FixedWidthParser()
        meta = p.extraire_metadata(path)
        assert meta["format"] == "fixedwidth_sage_pnm"
        assert meta["logiciel"] == "sage_pnm"
        assert meta["nb_lignes"] >= 5

    def test_metadata_error(self, tmp_path):
        """Covers lines 135-136: exception in metadata extraction."""
        p = FixedWidthParser()
        meta = p.extraire_metadata(Path("/nonexistent/file.pnm"))
        assert "erreur" in meta


class TestDetecterFormat:
    """Covers _detecter_format for both SAGE PNM and CIEL XIMPORT."""

    def test_detect_sage_pnm(self):
        p = FixedWidthParser()
        lines = [_sage_pnm_line() for _ in range(3)]
        assert p._detecter_format(lines) == "sage_pnm"

    def test_detect_ciel_ximport(self):
        p = FixedWidthParser()
        lines = [_ciel_ximport_line() for _ in range(3)]
        assert p._detecter_format(lines) == "ciel_ximport"

    def test_empty_lines(self):
        p = FixedWidthParser()
        assert p._detecter_format([]) is None

    def test_only_blank_lines(self):
        p = FixedWidthParser()
        assert p._detecter_format(["", "   ", "\n"]) is None

    def test_unrecognized_format(self):
        p = FixedWidthParser()
        assert p._detecter_format(["short line", "another"]) is None


class TestParserSagePnm:
    """Covers lines 201-202, 233-234, 260-318 (_parser_sage_pnm)."""

    def test_full_sage_pnm_parse(self, tmp_path):
        """Full SAGE PNM parsing with raison_sociale and journal tracking."""
        raison = "SOCIETE EXEMPLE SARL"
        lines = [
            raison,  # First line = company name (not a data line)
            _sage_pnm_line(journal="OD ", debit="    3500.00  ", credit="       0.00  "),
            _sage_pnm_line(journal="OD ", debit="       0.00  ", credit="    3500.00  "),
            _sage_pnm_line(journal="BQ ", debit="    1200.00  ", credit="       0.00  "),
        ]
        path = tmp_path / "sage.pnm"
        path.write_text("\n".join(lines), encoding="cp1252")

        p = FixedWidthParser()
        doc = _make_doc(name="sage.pnm", ftype=FileType.TEXTE, path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        decl = decls[0]
        assert decl.type_declaration == "export_comptable"
        assert decl.metadata["logiciel"] == "sage"
        assert decl.metadata["raison_sociale"] == raison
        assert "OD" in decl.metadata["journaux"]
        assert "BQ" in decl.metadata["journaux"]
        assert len(decl.cotisations) >= 2

    def test_sage_pnm_short_lines_skipped(self, tmp_path):
        """Lines shorter than 80 chars are skipped."""
        lines = [
            _sage_pnm_line(debit="    1000.00  ", credit="       0.00  "),
            "short line",
            _sage_pnm_line(debit="    2000.00  ", credit="       0.00  "),
        ]
        path = tmp_path / "sage_short.pnm"
        path.write_text("\n".join(lines), encoding="cp1252")

        p = FixedWidthParser()
        doc = _make_doc(name="sage_short.pnm", ftype=FileType.TEXTE, path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        # Short line should be skipped - only 2 cotisations
        assert len(decls[0].cotisations) == 2

    def test_sage_pnm_debit_and_credit(self, tmp_path):
        """Both debit and credit lines produce cotisations."""
        lines = [
            _sage_pnm_line(debit="    5000.00  ", credit="       0.00  "),
            _sage_pnm_line(debit="       0.00  ", credit="    2000.00  "),
        ]
        path = tmp_path / "sage_dc.pnm"
        path.write_text("\n".join(lines), encoding="cp1252")

        p = FixedWidthParser()
        doc = _make_doc(name="sage_dc.pnm", ftype=FileType.TEXTE, path=path)
        decls = p.parser(path, doc)
        cots = decls[0].cotisations
        assert len(cots) == 2
        # First: debit line
        assert cots[0].montant_patronal == Decimal("5000.00")
        # Second: credit line
        assert cots[1].montant_salarial == Decimal("2000.00")


class TestParserCielXimport:
    """Covers lines 260-318 (_parser_ciel_ximport)."""

    def test_full_ciel_parse(self, tmp_path):
        """Full CIEL XIMPORT parsing with D/C sens."""
        lines = [
            _ciel_ximport_line(n_mouvement="00001", journal="OD", montant="     3500.00 ", sens="D"),
            _ciel_ximport_line(n_mouvement="00002", journal="OD", montant="     2000.00 ", sens="C"),
            _ciel_ximport_line(n_mouvement="00003", journal="BQ", montant="     1500.00 ", sens="D"),
        ]
        path = tmp_path / "ciel.txt"
        path.write_text("\n".join(lines), encoding="cp1252")

        p = FixedWidthParser()
        doc = _make_doc(name="ciel.txt", ftype=FileType.TEXTE, path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        decl = decls[0]
        assert decl.metadata["logiciel"] == "ciel"
        assert "OD" in decl.metadata["journaux"]
        assert "BQ" in decl.metadata["journaux"]
        assert len(decl.cotisations) == 3

        # Check D/C sense
        cot_d = decl.cotisations[0]
        assert cot_d.montant_patronal == Decimal("3500.00")
        assert cot_d.montant_salarial == Decimal("0")

        cot_c = decl.cotisations[1]
        assert cot_c.montant_patronal == Decimal("0")
        assert cot_c.montant_salarial == Decimal("2000.00")

    def test_ciel_short_lines_skipped(self, tmp_path):
        lines = [
            _ciel_ximport_line(montant="     1000.00 ", sens="D"),
            "short",
            _ciel_ximport_line(montant="     2000.00 ", sens="C"),
        ]
        path = tmp_path / "ciel_short.txt"
        path.write_text("\n".join(lines), encoding="cp1252")

        p = FixedWidthParser()
        doc = _make_doc(name="ciel_short.txt", ftype=FileType.TEXTE, path=path)
        decls = p.parser(path, doc)
        assert len(decls[0].cotisations) == 2

    def test_ciel_zero_montant_skipped(self, tmp_path):
        lines = [
            _ciel_ximport_line(montant="        0.00 ", sens="D"),
            _ciel_ximport_line(montant="     5000.00 ", sens="D"),
        ]
        path = tmp_path / "ciel_zero.txt"
        path.write_text("\n".join(lines), encoding="cp1252")

        p = FixedWidthParser()
        doc = _make_doc(name="ciel_zero.txt", ftype=FileType.TEXTE, path=path)
        decls = p.parser(path, doc)
        # Zero amount should not produce a cotisation
        assert len(decls[0].cotisations) == 1


class TestFixedWidthParserMainParser:
    """Covers lines 144, 151 (parser method dispatch)."""

    def test_empty_file_raises(self, tmp_path):
        """Covers line 144: empty file."""
        path = tmp_path / "empty.pnm"
        path.write_text("a", encoding="cp1252")  # non-empty to pass _verifier_taille
        # But content will be single char - no recognized format
        p = FixedWidthParser()
        doc = _make_doc(name="empty.pnm", ftype=FileType.TEXTE, path=path)
        with pytest.raises(ParseError):
            p.parser(path, doc)

    def test_unrecognized_format_raises(self, tmp_path):
        path = tmp_path / "weird.pnm"
        path.write_text("This is not a fixed width format\nJust random text\n", encoding="cp1252")
        p = FixedWidthParser()
        doc = _make_doc(name="weird.pnm", ftype=FileType.TEXTE, path=path)
        with pytest.raises(ParseError, match="non reconnu"):
            p.parser(path, doc)

    def test_sage_format_detected_and_parsed(self, tmp_path):
        lines = [_sage_pnm_line(debit="    1000.00  ", credit="       0.00  ") for _ in range(3)]
        path = tmp_path / "sage.pnm"
        path.write_text("\n".join(lines), encoding="cp1252")
        p = FixedWidthParser()
        doc = _make_doc(name="sage.pnm", ftype=FileType.TEXTE, path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        assert decls[0].metadata["type_document"] == "export_comptable_sage_pnm"

    def test_ciel_format_detected_and_parsed(self, tmp_path):
        lines = [_ciel_ximport_line(montant="     1000.00 ", sens="D") for _ in range(3)]
        path = tmp_path / "ciel.txt"
        path.write_text("\n".join(lines), encoding="cp1252")
        p = FixedWidthParser()
        doc = _make_doc(name="ciel.txt", ftype=FileType.TEXTE, path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        assert decls[0].metadata["type_document"] == "export_comptable_ciel_ximport"


class TestLire:
    """Covers lines 326-328 (_lire with encoding fallback)."""

    def test_read_cp1252(self, tmp_path):
        path = tmp_path / "cp1252.txt"
        path.write_text("Soci\u00e9t\u00e9 Fran\u00e7aise", encoding="cp1252")
        content = FixedWidthParser._lire(path)
        assert "Soci" in content

    def test_read_utf8(self, tmp_path):
        path = tmp_path / "utf8.txt"
        path.write_text("Hello UTF-8 \u00e9\u00e0\u00fc", encoding="utf-8")
        content = FixedWidthParser._lire(path)
        assert "Hello" in content

    def test_read_fallback_latin1(self, tmp_path):
        """Covers lines 326-328: all encodings fail, fallback to latin-1 with errors=replace."""
        path = tmp_path / "binary.txt"
        # Write raw bytes that are valid in latin-1 but may cause issues
        path.write_bytes(b"Hello\x80\x81\x82World\n")
        content = FixedWidthParser._lire(path)
        assert "Hello" in content
        assert "World" in content


class TestExcelParserIntegration:
    """Integration tests with real Excel workbooks covering complex scenarios."""

    def test_multiple_sheets_some_valid(self, tmp_path):
        """Only sheets with valid data produce declarations."""
        path = tmp_path / "multi.xlsx"
        _write_excel(path, {
            "Valid": [
                ("Nom", "Prenom", "Salaire Brut"),
                ("DUPONT", "Jean", 3500),
            ],
            "Empty": [
                ("Random", "Headers"),
            ],
        })
        p = ExcelParser()
        doc = _make_doc(path=path)
        decls = p.parser(path, doc)
        # Only Valid sheet should produce a declaration
        assert len(decls) == 1

    def test_cotisation_linked_to_employee(self, tmp_path):
        """Cotisations are linked to their employee via employe_id."""
        path = tmp_path / "linked.xlsx"
        _write_excel(path, {
            "S1": [
                ("Nom", "Prenom", "Salaire Brut", "Montant Patronal"),
                ("DUPONT", "Jean", 3500, 525),
            ]
        })
        p = ExcelParser()
        doc = _make_doc(path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        assert len(decls[0].cotisations) == 1
        assert len(decls[0].employes) == 1
        # Cotisation should be linked to the employee
        assert decls[0].cotisations[0].employe_id == decls[0].employes[0].id

    def test_duplicate_employee_dedup_by_name(self, tmp_path):
        """Same employee on multiple rows is deduplicated."""
        path = tmp_path / "dedup.xlsx"
        _write_excel(path, {
            "S1": [
                ("Nom", "Prenom", "Salaire Brut", "Montant Patronal"),
                ("DUPONT", "Jean", 3500, 525),
                ("DUPONT", "Jean", 3500, 175),  # Same employee, different cotisation
            ]
        })
        p = ExcelParser()
        doc = _make_doc(path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        assert len(decls[0].employes) == 1  # Deduplicated
        assert len(decls[0].cotisations) == 2  # Both cotisations kept

    def test_sage_keyword_headers(self, tmp_path):
        """SAGE-specific header keywords are recognized."""
        path = tmp_path / "sage.xlsx"
        _write_excel(path, {
            "S1": [
                ("Nom de famille", "Prenom", "Brut soumis", "Mt Patronal"),
                ("MARTIN", "Sophie", 4200, 630),
            ]
        })
        p = ExcelParser()
        doc = _make_doc(path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        assert decls[0].employes[0].nom == "MARTIN"

    def test_header_on_row_3(self, tmp_path):
        """Header found on third row (after title rows)."""
        path = tmp_path / "offset.xlsx"
        _write_excel(path, {
            "S1": [
                ("SOCIETE ABC", None, None),
                ("Bulletin de Paie - Janvier 2024", None, None),
                ("Nom", "Prenom", "Salaire Brut"),
                ("DUPONT", "Jean", 3500),
            ]
        })
        p = ExcelParser()
        doc = _make_doc(path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        assert decls[0].employes[0].nom == "DUPONT"

    def test_row_with_all_mapped_values_empty(self, tmp_path):
        """Covers line 303: mapped_row ends up empty because values are all None/blank."""
        path = tmp_path / "emptyvals.xlsx"
        _write_excel(path, {
            "S1": [
                ("Nom", "Prenom", "Salaire Brut"),
                ("DUPONT", "Jean", 3500),
                ("", "", ""),  # Row with all empty strings -> mapped_row empty
                ("MARTIN", "Sophie", 4200),
            ]
        })
        p = ExcelParser()
        doc = _make_doc(path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        assert len(decls[0].employes) == 2

    def test_nir_dedup(self, tmp_path):
        """Employee deduplication by NIR."""
        path = tmp_path / "nir_dedup.xlsx"
        _write_excel(path, {
            "S1": [
                ("Nom", "Prenom", "NIR", "Salaire Brut"),
                ("DUPONT", "Jean", "1850175000001", 3500),
                ("DUPONT", "Jean", "1850175000001", 3500),  # duplicate
            ]
        })
        p = ExcelParser()
        doc = _make_doc(path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        assert len(decls[0].employes) == 1


# ============================================================
# Additional edge case tests for remaining coverage gaps
# ============================================================

class TestParserFeuilleEdgeCases:
    """Additional edge cases for _parser_feuille to cover remaining lines."""

    def test_mapped_row_all_whitespace_values(self, tmp_path):
        """Covers line 303: row where mapped columns exist but values are whitespace-only.
        The check `str(val).strip()` filters them out, leaving mapped_row empty."""
        path = tmp_path / "ws.xlsx"
        _write_excel(path, {
            "S1": [
                ("Nom", "Salaire Brut"),
                ("DUPONT", 3500),
                (" ", " "),  # whitespace-only values -> stripped to empty -> filtered out
            ]
        })
        p = ExcelParser()
        doc = _make_doc(path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        # Only 1 employee from first data row
        assert len(decls[0].employes) == 1

    def test_trouver_entete_fallback_skips_none_rows(self):
        """Covers line 393: None row in fallback loop of _trouver_entete."""
        p = ExcelParser()
        # No row has identity+amount -> falls to fallback loop
        # Fallback loop encounters None rows and skips them
        rows = [
            None,
            (),
            ("heures", "service"),  # mappable but no identity/amount
            (151, "Finance"),
        ]
        idx, header = p._trouver_entete(rows)
        # Should find row 2 in fallback
        assert idx == 2


class TestCielPartialFieldExtraction:
    """Covers lines 273-274 in fixedwidth_parser: partial field when start < len(ligne)."""

    def test_ciel_line_truncated_after_sens(self, tmp_path):
        """A CIEL line that is exactly 85 chars (minimum) - fields after sens are partial."""
        # Build a line that is exactly 85 chars - no pointage/analytique/libelle_compte
        line = _ciel_ximport_line()
        truncated = line[:86]  # Just past sens field at position 84
        lines = [truncated for _ in range(3)]
        path = tmp_path / "ciel_trunc.txt"
        path.write_text("\n".join(lines), encoding="cp1252")

        p = FixedWidthParser()
        doc = _make_doc(name="ciel_trunc.txt", ftype=FileType.TEXTE, path=path)
        decls = p.parser(path, doc)
        assert len(decls) == 1
        assert len(decls[0].cotisations) >= 1


class TestFixedWidthEmptyContent:
    """Covers line 144: parser with truly empty content after reading."""

    def test_parser_empty_content(self, tmp_path):
        """File has content but after split it has no recognizable format."""
        path = tmp_path / "onlynewlines.pnm"
        path.write_text("\n\n\n", encoding="cp1252")
        p = FixedWidthParser()
        doc = _make_doc(name="onlynewlines.pnm", ftype=FileType.TEXTE, path=path)
        with pytest.raises(ParseError):
            p.parser(path, doc)


class TestLireEncodingFallback:
    """Covers line 328: the ultimate fallback path in _lire."""

    def test_fallback_with_replacement_chars(self, tmp_path):
        """Write bytes that fail all standard encodings, requiring errors='replace'."""
        path = tmp_path / "tricky.txt"
        # cp1252, iso-8859-1, utf-8-sig, utf-8, latin-1 all handle single bytes,
        # so they won't fail. The fallback line 328 is only reached if all raise
        # UnicodeDecodeError. In practice, latin-1 accepts all byte values.
        # To cover line 326-327 (the continue), we mock read_text to fail on first attempts.
        p = FixedWidthParser()
        path.write_bytes(b"test content")
        call_count = 0
        original_read_text = Path.read_text

        def mock_read_text(self, encoding=None, errors=None):
            nonlocal call_count
            call_count += 1
            if errors != "replace":
                raise UnicodeDecodeError("test", b"", 0, 1, "forced error")
            return "test content"

        with patch.object(Path, "read_text", mock_read_text):
            content = FixedWidthParser._lire(path)
        assert content == "test content"
        # Should have tried multiple encodings then hit the fallback
        assert call_count >= 2
